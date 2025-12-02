import openpyxl

DIV_MO = 4.08 #Dividendo anual MO (ALTRIA GROUP, INC)
DIV_PF = 1.72 #Dividendo anual VZ (PFIZER INC)
DIV_VZ = 2.71 #Dividendo anual VZ (VERIZON COMMUNICATIONS INC.)

RAT_MO = 4 #Ratios de compra de acciones/CEDEAR MO
RAT_PF = 4 #Ratios de compra de acciones/CEDEAR PF
RAT_VZ = 4 #Ratios de compra de acciones/CEDEAR VZ

DIV_CD_MO = DIV_MO / 4 #Dividendo trimestral CEDEAR MO
DIV_CD_PF = DIV_PF / 4 #Dividendo trimestral CEDEAR
DIV_CD_VZ = DIV_VZ / 4 #Dividendo trimestral CEDEAR VZ

MO = 58.63 #Valor de la accion MO
PF = 24.24 #Valor de la accion PF
VZ = 43.27 #Valor de la accion VZ

CD_MO = 58.63 / RAT_MO #Valor del CEDEAR de MO
CD_PF = 24.24 / RAT_PF #Valor del CEDEAR de PF
CD_VZ = 43.27 / RAT_VZ #Valor del CEDEAR de VZ

ARS_USS = 1354.10 #Valor del dolar en pesos argentinos

periodo = 0 #Año inicial de inversion

INV_ARS = 300000 #Inversión mensual en pesos argentinos
INV_USS = INV_ARS / ARS_USS #Inversión mensual en dólares

cash= 0 #Dinero disponible en caja para invertir

ganancia_anual = 0 #Ganancia anual en pesos argentinos
cedears_MO = 0 #Cantidad de CEDEARs de MO
cedears_PF = 0 #Cantidad de CEDEARs de PF
cedears_VZ = 0 #Cantidad de CEDEARs de VZ

INV_TOTAL = 0 #Inversión total en pesos argentinos

ganancia_anual_requerida = 12000

TNA=0.08 #Tasa nominal anual de retorno esperada del 8%
TNM= TNA / 12 #Tasa nominal mensual de retorno esperada del 8%

ganancias_totales=0 #Ganancias totales en dolares

diccionario_dividendos_con_cedears={
    "Valor CEDEAR MO (U$S)": CD_MO,
    "Valor CEDEAR PF (U$S)": CD_PF,
    "Valor CEDEAR VZ (U$S)": CD_VZ,
    "Valor dividendo trimestral MO (U$S)": DIV_CD_MO,
    "Valor dividendo trimestral PF (U$S)": DIV_CD_PF,
    "Valor dividendo trimestral VZ (U$S)": DIV_CD_VZ,
    "Tipo de cambio AR$/U$S": ARS_USS,
    "Inversión Mensual en pesos": INV_ARS,
    "Inversión Mensual en dolares": INV_USS,
    "Ganancia Anual Requerida (U$S)": ganancia_anual_requerida,
    "TNA (%)": TNA * 100,
    "TNM (%)": TNM * 100,
    "Matriz de inversión": []
}



#Año 16; Ganancia anual: 1263.01 U$S; Inversión total: 20.400.000.00 ARS; 
# CEDEARs MO: 94.0; CEDEARs PF: 336.0; CEDEARs VZ: 136.0

#ganancia_anual<1200 3 años 100 // INV = 100000
#ganancia_anual<2400 4 años 200 // INV = 100000
#ganancia_anual<3600 4 años 300 // INV = 100000
#ganancia_anual<4800 5 años 400 // INV = 100000
#ganancia_anual<6000 6 años 500 // INV = 100000
#ganancia_anual<7200 6 años 600 // INV = 100000
#ganancia_anual<8400 6 años 700 // INV = 100000
#ganancia_anual<9600 7 años 800 // INV = 100000
#ganancia_anual<10800 7 años 900 // INV = 100000
#ganancia_anual<12000 7 años 1000 // INV = 100000

#ganancia_anual<1200 2 años 100 // INV = 200000
#ganancia_anual<2400 3 años 200 // INV = 200000
#ganancia_anual<3600 3 años 300 // INV = 200000
#ganancia_anual<4800 4 años 400 // INV = 200000
#ganancia_anual<6000 4 años 500 // INV = 200000
#ganancia_anual<7200 4 años 600 // INV = 200000
#ganancia_anual<8400 5 años 700 // INV = 200000
#ganancia_anual<9600 5 años 800 // INV = 200000
#ganancia_anual<10800 5 años 900 // INV = 200000
#ganancia_anual<12000 5 años 1000 // INV = 200000

#ganancia_anual<1200 1 años 100 // INV = 300000
#ganancia_anual<2400 2 años 200 // INV = 300000
#ganancia_anual<3600 3 años 300 // INV = 300000
#ganancia_anual<4800 3 años 400 // INV = 300000
#ganancia_anual<6000 3 años 500 // INV = 300000
#ganancia_anual<7200 3 años 600 // INV = 300000
#ganancia_anual<8400 4 años 700 // INV = 300000
#ganancia_anual<9600 4 años 800 // INV = 300000
#ganancia_anual<10800 4 años 900 // INV = 300000
#ganancia_anual<12000 4 años 1000 // INV = 300000

while (ganancia_anual * TNA) < ganancia_anual_requerida:
    for i in range(1,13):
        INV_TOTAL += INV_ARS
        inv_mensual = INV_USS
        if i==1 or i==4 or i==7 or i==10: #Meses con dividendos de MO
            cedears_VZ += inv_mensual // CD_VZ #Comprar CEDEARs de VZ
            ganancia_mensual = cedears_MO * DIV_CD_MO *(1 - 0.3) * (1 - 0.06 * (1 + 0.21))
            resto= inv_mensual / CD_VZ - inv_mensual // CD_VZ
            cash = resto + ganancia_mensual + cash * (1 + TNM) # Reinvertir el dinero en caja del mes anterior
            ganancia_anual += cash
        elif i==2 or i==5 or i==8 or i==11: #Meses con dividendos de PF
            cedears_MO += inv_mensual // CD_MO #Comprar CEDEARs de MO
            ganancia_mensual = cedears_PF * DIV_CD_PF *(1 - 0.3) * (1 - 0.06 * (1 + 0.21))
            resto = inv_mensual / CD_MO - inv_mensual // CD_MO
            cash = resto + ganancia_mensual + cash * (1 + TNM) # Reinvertir el dinero en caja del mes anterior
            ganancia_anual += cash
        elif i==3 or i==6 or i==9 or i==12: #Meses con dividendos de VZ
            cedears_PF += inv_mensual // CD_PF #Comprar CEDEARs de PF
            ganancia_mensual = cedears_VZ * DIV_CD_VZ *(1 - 0.3) * (1 - 0.06 * (1 + 0.21))
            resto = inv_mensual / CD_PF - inv_mensual // CD_PF
            cash = resto + ganancia_mensual + cash * (1 + TNM) # Reinvertir el dinero en caja del mes anterior
            ganancia_anual += cash
        else:
            print("Error en el mes: ", i)
            break
        if i==12: #Al final del año, reiniciar el ciclo
            ganancias_totales += ganancia_anual

        diccionario_dividendos_con_cedears["Matriz de inversión"].append({
        "Año": periodo,
        "Mes": i,
        "Inversión Mensual (ARS)": INV_ARS,
        "Inversión Mensual (USD)": INV_USS,
        "CEDEARs MO": cedears_MO,
        "CEDEARs PF": cedears_PF,
        "CEDEARs VZ": cedears_VZ,
        "Ganancia Mensual (USD)": ganancia_mensual,
        "Dinero en Caja (USD)": cash,
        "Ganancia Anual (USD)": ganancia_anual,
        "Inversión Total (ARS)": INV_TOTAL,
        "cash": cash,
        "ganancia_total": ganancia_anual
        })

        if i==12:
            print(f"Año {periodo}; Ganancia anual: {ganancia_anual:.2f} U$S; Inversión total: {INV_TOTAL:.2f} ARS; CEDEARs MO: {cedears_MO}; CEDEARs PF: {cedears_PF}; CEDEARs VZ: {cedears_VZ}")
            periodo += 1

        ganancia_mensual = 0 #Reiniciar ganancia mensual al final de cada mes

print("Fin del ciclo de inversión con dividendos reinvertidos.")
#print(diccionario_dividendos_con_cedears)

# Crear un libro y una hoja
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Resumen"

# Escribir valores específicos en celdas concretas
ws["A1"] = "Valor CEDEAR MO (U$S)"
ws["B1"] = diccionario_dividendos_con_cedears["Valor CEDEAR MO (U$S)"]

ws["A2"] = "Valor CEDEAR PF (U$S)"
ws["B2"] = diccionario_dividendos_con_cedears["Valor CEDEAR PF (U$S)"]

ws["A3"] = "Valor CEDEAR VZ (U$S)"
ws["B3"] = diccionario_dividendos_con_cedears["Valor CEDEAR VZ (U$S)"]

ws["A4"] = "Tipo de cambio AR$/U$S"
ws["B4"] = diccionario_dividendos_con_cedears["Tipo de cambio AR$/U$S"]

ws["A5"] = "Inversión Mensual en pesos"
ws["B5"] = diccionario_dividendos_con_cedears["Inversión Mensual en pesos"]

ws["A6"] = "Inversión Mensual en dolares"
ws["B6"] = diccionario_dividendos_con_cedears["Inversión Mensual en dolares"]

ws["A7"] = "Ganancia Anual Requerida (U$S)"
ws["B7"] = diccionario_dividendos_con_cedears["Ganancia Anual Requerida (U$S)"]

ws["A8"] = "Valor dividendo trimestral MO (U$S)"
ws["B8"] = diccionario_dividendos_con_cedears["Valor dividendo trimestral MO (U$S)"]

ws["A9"] = "Valor dividendo trimestral PF (U$S)"
ws["B9"] = diccionario_dividendos_con_cedears["Valor dividendo trimestral PF (U$S)"]

ws["A10"] = "Valor dividendo trimestral VZ (U$S)"
ws["B10"] = diccionario_dividendos_con_cedears["Valor dividendo trimestral VZ (U$S)"]

# Si quieres agregar la matriz en otra hoja:
ws2 = wb.create_sheet("Matriz de inversión")
matriz = diccionario_dividendos_con_cedears["Matriz de inversión"]

# Escribir encabezados
if matriz:
    encabezados = list(matriz[0].keys())
    ws2.append(encabezados)
    for fila in matriz:
        ws2.append(list(fila.values()))

# Guardar el archivo
wb.save(f"reporte_inversion_dividendos_FCI_TNA_{int(TNA*100)}%.xlsx")
print("Datos exportados a reporte_inversion_dividendos_FCI.xlsx")